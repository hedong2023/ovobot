from openpyxl import Workbook
from openpyxl import load_workbook
import argparse
import re

fpLib = {
    "soic-8": [
        "sop8",
        "sop-8",
    ],
    "tssop-16": [
        "tssop16"
    ],
    "sod-123": [
        "sod123"
    ]
}

def convertRefToList(ref):
    str1 = ref.replace(' ', '')

    if str1[-1] == '\n':
        str1 = str1[:-1]

    if str1[-1] == ',' or str1[-1] == '，':
        str1 = str1[:-1]
    
    refList = re.split(",|，", str1)

    return refList

def removeChinese(val):
    pattern = re.compile(r'[\u4e00-\u9fa5]')
    result = re.sub(pattern, '', val)

    return result

def convertValue(val):
    str1 = val.replace(' ', '')
    result = removeChinese(str1)

    return result.lower()

def convertFootprint(category, footprint):
    if footprint != None:
        fp = re.split(":", footprint)[1]

        conn_regex = re.compile(r'Conn_.*')
        
        if category == 'C' or category == 'R' or category == 'D' or category == 'D_Zener' or category == 'Fuse':
            fp_short = re.split("_", fp)[1]
        elif category == 'C_Polarized':
            fp_short = re.split("_", fp)[2]
        elif category == 'Buzzer':
            fp_short = re.split("RM", re.split("_", fp)[1])[0]
        elif category == 'INDUCTOR':
            fp_short = re.split("_", fp)[2]
        elif category == 'L_Coupled_1213':
            fp_short = re.split("_", fp)[2]
        elif re.match(conn_regex, category):
            fp_short = ''
        else:
            fp_short = re.split("_", fp)[0]
    else:
        fp_short = ''

    # print(fp_short)

    return fp_short


parser = argparse.ArgumentParser(description='Process some integers.')
parser.add_argument('file1', metavar='file1', nargs=1,
                    help='BOM原文件')
parser.add_argument('file2', metavar='file2', nargs=1,
                    help='odoo产品目录')

args = parser.parse_args()

wb_source = load_workbook(args.file1[0])
ws_source = wb_source.active

wb_products = load_workbook(args.file2[0])
ws_products = wb_products.active

for row_source in ws_source.iter_rows(min_row=7, values_only=True):
    if row_source[2] != None:
        text =  row_source[2].replace(')', '')
        vals = re.split("/|\(|@", text)

        fp = convertFootprint(row_source[3], row_source[4])

        if fp != '':
            vals.append(fp)

        # print(vals)
        find = 0
        for row_products in ws_products.iter_rows(min_row=2, values_only=True):
            odoo_value_conv = convertValue(row_products[1])
            # print(odoo_value_conv)

            valMatch = 1
            for val in vals:
                if val.lower() not in odoo_value_conv:
                    valMatch = 0

                    if val.lower() in fpLib.keys():
                        for item in fpLib[val.lower()]:
                            if item in odoo_value_conv:
                                valMatch = 1

                    if valMatch == 0:
                        break

            if valMatch == 1:
                print("原始参数：%s 匹配参数：%s"%(row_source[2] + ' ' + fp, row_products[1]))
                find = 1
                break

        if find == 0:
            print("未匹配，原始参数：%s"%(row_source[2] + ' ' + fp))
