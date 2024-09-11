from openpyxl import Workbook
from openpyxl import load_workbook
import argparse
import re

fpLib = {
    "soic-8": [
        "sop8",
        "sop-8",
    ],
    "tssop-16": [
        "tssop16"
    ]
}

def convertRefToList(ref):
    str1 = ref.replace(' ', '')

    if str1[-1] == '\n':
        str1 = str1[:-1]

    if str1[-1] == ',' or str1[-1] == '，':
        str1 = str1[:-1]
    
    refList = re.split(",|，", str1)

    return refList

def removeChinese(val):
    pattern = re.compile(r'[\u4e00-\u9fa5]')
    result = re.sub(pattern, '', val)

    return result

def convertValue(val):
    str1 = val.replace(' ', '')
    result = removeChinese(str1)

    return result.lower()

def convertFootprint(category, footprint):
    fp = re.split(":", footprint)[1]

    conn_regex = re.compile(r'Conn_.*')
    
    if category == 'C' or category == 'R' or category == 'D' or category == 'D_Zener' or category == 'Fuse':
        fp_short = re.split("_", fp)[1]
    elif category == 'C_Polarized':
        fp_short = re.split("_", fp)[2]
    elif category == 'Buzzer':
        fp_short = re.split("RM", re.split("_", fp)[1])[0]
    elif category == 'INDUCTOR':
        fp_short = re.split("_", fp)[2]
    elif category == 'L_Coupled_1213':
        fp_short = re.split("_", fp)[2]
    elif re.match(conn_regex, category):
        fp_short = ''
    else:
        fp_short = re.split("_", fp)[0]

    # print(fp_short)

    return fp_short


parser = argparse.ArgumentParser(description='Process some integers.')
parser.add_argument('file1', metavar='file1', nargs=1,
                    help='odoo下载的BOM')
parser.add_argument('row_start_odoo', metavar='N', type=int, nargs=1,
                    help='开始比较的行')
parser.add_argument('file2', metavar='file2', nargs=1,
                    help='用于比较的原始BOM')
parser.add_argument('row_start_source', metavar='N', type=int, nargs=1,
                    help='开始比较的行')

args = parser.parse_args()

wb_odoo = load_workbook(args.file1[0])
ws_odoo = wb_odoo.active

wb_source = load_workbook(args.file2[0])
ws_source = wb_source.active

ws_source_len = ws_source.max_row
source_match = []

for x in range(ws_source_len):
    source_match.append(0)

for row_odoo in ws_odoo.iter_rows(min_row=args.row_start_odoo[0], values_only=True):
    if row_odoo[4] != None:
        list1 = convertRefToList(row_odoo[4])
        list1.sort()
        i = args.row_start_source[0]

        for row_source in ws_source.iter_rows(min_row=args.row_start_source[0], values_only=True):
            if row_source[0] != None:
                list2 = convertRefToList(row_source[0])
                list2.sort()

                if list1 == list2:
                    if str(row_odoo[3]) != str(row_source[1]):
                        print("以下行数量不匹配：")
                        print(row_odoo)
                    else:
                        odoo_value_conv = convertValue(row_odoo[1])
                        # print(odoo_value_conv)
                        text =  row_source[2].replace(')', '')
                        vals = re.split("/|\(|@", text)

                        fp = convertFootprint(row_source[3], row_source[4])

                        if fp != '':
                            vals.append(fp)

                        # print(vals)
                    
                        for val in vals:
                            if val.lower() not in odoo_value_conv:
                                find = 0

                                if val.lower() in fpLib.keys():
                                    for item in fpLib[val.lower()]:
                                        if item in odoo_value_conv:
                                            find = 1
                                
                                if find == 0:
                                    print("以下参数不匹配：%s odoo参数: %s, 原始参数: %s" % (val, row_odoo[1], row_source[2] + ' ' + row_source[4]))
                                
                                break

                    source_match[i-1] = 1

                    break
            i += 1
        
        if (i > ws_source_len):
            print("以下行未匹配到：")
            print(row_odoo)
    else:
        print("以下行位号为空：")
        print(row_odoo)

print("\n\r\n\r以下是BOM原文件未匹配到的行：")
i = args.row_start_source[0]

for row_source in ws_source.iter_rows(min_row=args.row_start_source[0], values_only=True):
    if source_match[i - 1] != 1 and row_source[0] != None:
        print(row_source)
    i += 1